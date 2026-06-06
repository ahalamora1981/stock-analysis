import io
from datetime import date

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.models.stock import Stock
from app.models.analysis import CompositeScore
from app.models.position import Position

router = APIRouter(prefix="/export", tags=["export"])


@router.get("/excel")
async def export_excel(db: AsyncSession = Depends(get_db)):
    """Export analysis report as Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    # Sheet 1: Stock Analysis
    ws1 = wb.active
    ws1.title = "股票分析"
    headers = ["排名", "代码", "名称", "综合评分", "估值分", "技术面分", "基本面分", "资金流分", "动量分", "等级"]
    ws1.append(headers)

    scores = (await db.execute(
        select(CompositeScore, Stock.code, Stock.name)
        .join(Stock, CompositeScore.stock_id == Stock.id)
        .order_by(CompositeScore.rank)
    )).all()

    grade_map = {1: "优秀", 2: "良好", 3: "一般", 4: "较差"}
    for cs, code, name in scores:
        ws1.append([
            cs.rank, code, name,
            round(cs.total_score, 1),
            round(cs.valuation_score, 1),
            round(cs.technical_score, 1),
            round(cs.fundamental_score, 1),
            round(cs.capital_flow_score, 1),
            round(cs.momentum_score, 1),
            grade_map.get(cs.grade, "--"),
        ])

    # Style headers
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Sheet 2: Positions
    ws2 = wb.create_sheet("持仓")
    ws2.append(["代码", "名称", "持仓数量", "成本价", "最新价", "盈亏金额", "盈亏比例"])

    positions = (await db.execute(
        select(Position, Stock.code, Stock.name)
        .join(Stock, Position.stock_id == Stock.id)
        .where(Position.total_shares > 0)
    )).all()

    for pos, code, name in positions:
        from app.models.daily_data import StockDailyData
        latest = (await db.execute(
            select(StockDailyData)
            .where(StockDailyData.stock_id == pos.stock_id)
            .order_by(StockDailyData.date.desc())
            .limit(1)
        )).scalar_one_or_none()
        price = latest.close if latest else 0
        pnl = (price - pos.avg_cost) * pos.total_shares
        pnl_pct = (pnl / pos.total_cost * 100) if pos.total_cost > 0 else 0
        ws2.append([code, name, pos.total_shares, round(pos.avg_cost, 2),
                     round(price, 2), round(pnl, 2), f"{pnl_pct:.2f}%"])

    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = date.today().isoformat()
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=stock_analysis_{today}.xlsx"},
    )
